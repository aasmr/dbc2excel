'''
Created on 1 февр. 2019 г.

@author: Инженер МГТУ им Н. Э. Баумана СМ10 Смирнов А. А. (мл.)
'''

import openpyxl, sys, re
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
from PySide2 import QtWidgets

def log_uncaught_exceptions(ex_cls, ex, tb):
    text = '{}: {}:\n'.format(ex_cls.__name__, ex)
    import traceback
    text += ''.join(traceback.format_tb(tb))
    print(text)
    QtWidgets.QMessageBox.critical(None, 'Error', text)
    #quit()
    
''' 
    openpyxl - работа с Excel;
    sys - не помню зачем, но надо
    re - работа с регулярными выражениями
    openpyxl.styles - позволяет настроить шрифты и выравнивание
    PySide2 - аналог PyQt с коммерческими возможностями
'''

class Main(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        width=240
        height=128
        #q=QtWidgets.QDesktopWidget().availableGeometry()
        #x=(q.width()-width)/2
        #y=(q.height()-height)/2
        
        self.setObjectName('MainWindow')
        self.setWindowTitle('dbc2excel')
        self.resize(width, height)
        #self.move(x, y)
        
        new_btn=QtWidgets.QPushButton('Создать новую таблицу', self)
        add_sheet_btn=QtWidgets.QPushButton('Добавить новый лист', self)
        add_sig_btn=QtWidgets.QPushButton('Обновить таблицу сигналов', self)
        exit_btn=QtWidgets.QPushButton('Выход', self)
          
        vbox=QtWidgets.QVBoxLayout(self)
        vbox.addWidget(new_btn)
        vbox.addWidget(add_sheet_btn)
        vbox.addWidget(add_sig_btn)
        vbox.addWidget(exit_btn)
        
        new_btn.clicked.connect(self.new_excel_win)
        add_sheet_btn.clicked.connect(self.add_sheet_win)
        add_sig_btn.clicked.connect(self.add_sig_win)
        exit_btn.clicked.connect(self.exit)
    def exit(self):
        self.close()    
    def new_excel_win(self):              
        class NewExcelDiag (QtWidgets.QWidget):
            def __init__(self):
                super().__init__()
                self.initUI()
                
            def initUI(self):
                width=640
                height=240
                q=QtWidgets.QDesktopWidget().availableGeometry()
                x=(q.width()-width)/2
                y=(q.height()-height)/2
                
                self.setObjectName('NewExcelDialog')
                self.setWindowTitle('Создать новую таблицу')
                self.resize(width, height)
                self.move(x, y)
                
                self.pathDBC=QtWidgets.QLineEdit(self)
                pathDBC_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathDBC_label=QtWidgets.QLabel(self)
                pathDBC_label.setText('Введите путь к DBC:')
                
                self.pathExcel=QtWidgets.QLineEdit(self)
                pathExcel_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathExcel_label=QtWidgets.QLabel(self)
                pathExcel_label.setText('Где сохранить таблицу Excel?')
                
                self.nameExcel=QtWidgets.QLineEdit(self)
                nameExcel_label=QtWidgets.QLabel(self)
                nameExcel_label.setText('Введите имя таблицы:')
                
                self.nameList=QtWidgets.QLineEdit(self)
                nameList_label=QtWidgets.QLabel(self)
                nameList_label.setText('Введите название листа:')
                
                next_btn=QtWidgets.QPushButton('Далее', self)
                
                gridbox=QtWidgets.QGridLayout(self)
                
                gridbox.addWidget(pathDBC_label, 0, 0)
                gridbox.addWidget(self.pathDBC, 1, 0)
                gridbox.addWidget(pathDBC_Viewbtn, 1, 1)
                
                gridbox.addWidget(pathExcel_label, 2, 0)
                gridbox.addWidget(self.pathExcel, 3, 0)
                gridbox.addWidget(pathExcel_Viewbtn, 3, 1)
                
                gridbox.addWidget(nameExcel_label, 4, 0)
                gridbox.addWidget(self.nameExcel, 5, 0, 1, 2)
                
                gridbox.addWidget(nameList_label, 6, 0)
                gridbox.addWidget(self.nameList, 7, 0, 1, 2)
                
                gridbox.addWidget(next_btn, 8, 1)
                
                pathDBC_Viewbtn.clicked.connect(self.openDBCpath)
                pathExcel_Viewbtn.clicked.connect(self.openExcelpath)
                next_btn.clicked.connect(self.next)
                
            def openDBCpath(self):
                dbcpath=QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '*.dbc')[0]
                self.pathDBC.setText(dbcpath)
            
            def openExcelpath(self):
                excelpath=QtWidgets.QFileDialog.getExistingDirectory(self, 'Save Excel', './')
                self.pathExcel.setText(excelpath)
                
            def next(self):
                """
                Обработать, когда ничего не введено
                Обработать удачное завершение
                """
                
                dbcpath=self.pathDBC.text()
                Excelpath=self.pathExcel.text()
                Excelname=self.nameExcel.text()
                Listname=self.nameList.text()
                def err():
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Critical)
                    msg.setText("Неверный ввод данных")
                    msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Error!")
                    msg.show()
                if dbcpath == '':
                    err()
                elif Excelpath == '':
                    err()
                elif Excelname == '':
                    err()
                elif Listname == '':
                    err()   
                else:
                    Excelpath=Excelpath+'/'+Excelname+'.xlsx'
                    new_excel(Excelpath, dbcpath, Listname)
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Information)
                    msg.setText("Создание таблицы произошло успешно!")
#msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Успех!")
                    msg.show()
                
                                
        self.window=NewExcelDiag()
        self.window.show()
        
    def add_sheet_win(self):              
        class AddSheetDiag (QtWidgets.QWidget):
            def __init__(self):
                super().__init__()
                self.initUI()
                
            def initUI(self):
                width=640
                height=180
                q=QtWidgets.QDesktopWidget().availableGeometry()
                x=(q.width()-width)/2
                y=(q.height()-height)/2
                
                self.setObjectName('AddSheetDialog')
                self.setWindowTitle('Создать новый лист')
                self.resize(width, height)
                self.move(x, y)
                
                self.pathDBC=QtWidgets.QLineEdit(self)
                pathDBC_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathDBC_label=QtWidgets.QLabel(self)
                pathDBC_label.setText('Введите путь к DBC:')
                
                self.pathExcel=QtWidgets.QLineEdit(self)
                pathExcel_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathExcel_label=QtWidgets.QLabel(self)
                pathExcel_label.setText('Введите путь к таблице Excel:')
                               
                self.nameList=QtWidgets.QLineEdit(self)
                nameList_label=QtWidgets.QLabel(self)
                nameList_label.setText('Введите название листа:')
                
                next_btn=QtWidgets.QPushButton('Далее', self)
                
                gridbox=QtWidgets.QGridLayout(self)
                
                gridbox.addWidget(pathDBC_label, 0, 0)
                gridbox.addWidget(self.pathDBC, 1, 0)
                gridbox.addWidget(pathDBC_Viewbtn, 1, 1)
                
                gridbox.addWidget(pathExcel_label, 2, 0)
                gridbox.addWidget(self.pathExcel, 3, 0)
                gridbox.addWidget(pathExcel_Viewbtn, 3, 1)
                                
                gridbox.addWidget(nameList_label, 4, 0)
                gridbox.addWidget(self.nameList, 5, 0, 1, 2)
                
                gridbox.addWidget(next_btn, 6, 1)
                
                pathDBC_Viewbtn.clicked.connect(self.openDBCpath)
                pathExcel_Viewbtn.clicked.connect(self.openExcelpath)
                next_btn.clicked.connect(self.next)
                
            def openDBCpath(self):
                dbcpath=QtWidgets.QFileDialog.getOpenFileName(self, 'Open DBC file', './', '*.dbc')[0]
                self.pathDBC.setText(dbcpath)
            
            def openExcelpath(self):
                excelpath=QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel file', './', '*.xlsx')[0]
                self.pathExcel.setText(excelpath)
                
            def next(self):
                """
                Обработать, когда ничего не введено
                Обработать удачное завершение
                """
                
                dbcpath=self.pathDBC.text()
                Excelpath=self.pathExcel.text()
                Listname=self.nameList.text()
                def err():
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Critical)
                    msg.setText("Неверный ввод данных")
                    msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Error!")
                    msg.show()
                if dbcpath == '':
                    err()
                elif Excelpath == '':
                    err()
                elif Listname == '':
                    err()   
                else:
                    add_sheet(Excelpath, dbcpath, Listname)
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Information)
                    msg.setText("Добавление листа произошло успешно!")
#msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Успех!")
                    msg.show()
                                         
        self.window=AddSheetDiag()
        self.window.show()

    def add_sig_win(self):
        class AddSigDiag(QtWidgets.QWidget):
            def __init__(self):
                super().__init__()
                self.initUI()
            
            def initUI(self):
                width=640
                height=180
                q=QtWidgets.QDesktopWidget().availableGeometry()
                x=(q.width()-width)/2
                y=(q.height()-height)/2
                
                self.setObjectName('AddSignalDialog')
                self.setWindowTitle('Обновить таблицу сигналов')
                self.resize(width, height)
                self.move(x, y)
                
                self.pathDBC=QtWidgets.QLineEdit(self)
                pathDBC_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathDBC_label=QtWidgets.QLabel(self)
                pathDBC_label.setText('Введите путь к DBC:')
                
                self.pathExcel=QtWidgets.QLineEdit(self)
                pathExcel_Viewbtn=QtWidgets.QPushButton('Обзор...', self)
                pathExcel_label=QtWidgets.QLabel(self)
                pathExcel_label.setText('Введите путь к таблице Excel:')
                               
                self.nameList=QtWidgets.QComboBox(self)
                nameList_label=QtWidgets.QLabel(self)
                nameList_label.setText('Выберите лист:')
                
                next_btn=QtWidgets.QPushButton('Далее', self)
                
                gridbox=QtWidgets.QGridLayout(self)
                
                gridbox.addWidget(pathDBC_label, 0, 0)
                gridbox.addWidget(self.pathDBC, 1, 0)
                gridbox.addWidget(pathDBC_Viewbtn, 1, 1)
                
                gridbox.addWidget(pathExcel_label, 2, 0)
                gridbox.addWidget(self.pathExcel, 3, 0)
                gridbox.addWidget(pathExcel_Viewbtn, 3, 1)
                                
                gridbox.addWidget(nameList_label, 4, 0)
                gridbox.addWidget(self.nameList, 5, 0, 1, 2)
                
                gridbox.addWidget(next_btn, 6, 1)
                
                pathDBC_Viewbtn.clicked.connect(self.openDBCpath)
                pathExcel_Viewbtn.clicked.connect(self.openExcelpath)
                next_btn.clicked.connect(self.next)
                
            def openDBCpath(self):
                dbcpath=QtWidgets.QFileDialog.getOpenFileName(self, 'Open DBC file', './', '*.dbc')[0]
                self.pathDBC.setText(dbcpath)
            
            def openExcelpath(self):
                excelpath=QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel file', './', '*.xlsx')[0]
                self.pathExcel.setText(excelpath)
                wb=openpyxl.load_workbook(excelpath, data_only=True)
                sheet_names=wb.sheetnames
                self.nameList.clear()
                for i in sheet_names:
                    self.nameList.addItem(i)
                
            def next(self):
                """
                Обработать, когда ничего не введено
                Обработать удачное завершение
                """
                
                dbcpath=self.pathDBC.text()
                Excelpath=self.pathExcel.text()
                Listname=self.nameList.currentText()
                
                def err():
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Critical)
                    msg.setText("Неверный ввод данных")
                    msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Error!")
                    msg.show()
                if dbcpath == '':
                    err()
                elif Excelpath == '':
                    err()
                elif Listname == '':
                    err()
                else:
                    add_sig(Excelpath, dbcpath, Listname)
                    msg = QtWidgets.QMessageBox(self)
                    msg.setIcon(QtWidgets.QMessageBox.Information)
                    msg.setText("Изменение сигнала произошло успешно!")
#msg.setInformativeText("Заполните все строки!")
                    msg.setWindowTitle("Успех!")
                    msg.show()
                
                
                
                                
        self.window=AddSigDiag()
        self.window.show()

class DBC():
    def __init__(self):
        pass
    class message(): #Подкласс записи сообщений
        def __init__(self, ID, name, signals):
            self.ID=ID
            self.name=name
            self.signals=signals
    class signal(): #Подкласс записи сигналов
        def __init__(self, name, startbit, length, factor, offset, unit, sig_com, sig_values, sig_range):
            self.name=name
            self.startbit=startbit
            self.length=length
            self.factor=factor
            self.offset=offset
            self.unit=unit
            self.sig_com=sig_com
            self.sig_values=sig_values
            self.sig_range=sig_range

def DBC_open(path): #функция открытия файла .dbc и записи информации в класс
    file=open(path, 'r')
    string=file.readline()
    dbc=[]
    while string!='':
        if string[0:3] == 'BO_':

            #ID сообщения
            message_id=''
            i=4
            char=string[i]
            while char != ' ':
                message_id=message_id+char
                i=i+1
                char=string[i]
            message_id=hex(int('0'+f"{int(message_id):b}"[1:], 2))[2:].upper()
            while len(message_id) !=8:
                message_id='0'+message_id
            #print(message_id)

            #Название сообщения
            message_name=''
            i=i+1
            char=string[i]
            while char != ':':
                message_name=message_name+char
                i=i+1
                char=string[i]
            #print(message_name)
                
            string=file.readline()
            signals=[]
            while string !='\n':
                if string[1:4] == 'SG_':

                    #Название сигнала
                    signal_name=''
                    i=5
                    char=string[i]
                    while char!=' ':
                        signal_name=signal_name+char
                        i=i+1
                        char=string[i]
                    #print(signal_name)

                    #Начальный бит сигнала
                    signal_startbit=''
                    char=string[i]
                    while char!=':':
                        i=i+1
                        char=string[i]
                    i=i+2
                    char=string[i]
                    while char!='|':
                        signal_startbit=signal_startbit+char
                        i=i+1
                        char=string[i]
                    #print(string)
                    #print(signal_startbit)

                    #Длина сигнала в битах
                    signal_length=''
                    i=i+1
                    char=string[i]
                    while char!='@':
                        signal_length=signal_length+char
                        i=i+1
                        char=string[i]
                    #print(signal_length)

                    #Коэффициент, на которое умнажается значение сигнала
                    signal_factor=''
                    i=i+5
                    char=string[i]
                    while char!=',':
                        signal_factor=signal_factor+char
                        i=i+1
                        char=string[i]
                    #print(signal_factor)

                    #Смещение, которое прибавляется к значению сигнала
                    signal_offset=''
                    i=i+1
                    char=string[i]
                    while char!=')':
                        signal_offset=signal_offset+char
                        i=i+1
                        char=string[i]
                    #print(signal_offset)
                    
                    #Единицы измерения значений сигнала
                    signal_unit=''
                    while string[i] != '"':
                        i=i+1
                    i=i+1
                    char=string[i]
                    while char!='"':
                        signal_unit=signal_unit+char
                        i=i+1
                        char=string[i]
                    #print(signal_unit)
                    
                    #Пределы измерения
                    signal_range=''
                    result=re.findall(r'(\[.+\])', string)
                    result=re.findall(r'[^\[\]\|]+|[^\[\]\|]+', result[0])
                    if result[0] != result[1]:
                        signal_range=result[0]+' до '+result[1]
                    
                    ukz=file.tell()
                    string=file.readline()
                    sig_com=''
                    sig_value=''
                    
                    while string!='':
                        if string[0:7] == 'CM_ SG_':
                            #ID сообщения
                            mes_id=''
                            i=8
                            char=string[i]
                            while char != ' ':
                                mes_id=mes_id+char
                                i=i+1
                                char=string[i]
                            mes_id=hex(int('0'+f"{int(mes_id):b}"[1:], 2))[2:].upper()
                            while len(mes_id) !=8:
                                mes_id='0'+mes_id
                                
                            if mes_id == message_id:
                                i=i+1
                                char=string[i]
                                sig_name=''
                                while char != ' ':
                                    sig_name=sig_name+char
                                    i=i+1
                                    char=string[i]
                                
                                if sig_name == signal_name:
                                    i=i+1
                                    char=string[i]
                                    sig_com=''
                                    try:
                                        char_check=string[i]+string[i+1]+string[i+2]
                                    except:
                                        char_check=''
                                    while char_check != '";\n':
                                        sig_com=sig_com+string[i]
                                        i=i+1
                                        if char =='\n':
                                            string=file.readline()
                                            i=0
                                        char=string[i]
                                        try:
                                            char_check=string[i]+string[i+1]+string[i+2]
                                        except:
                                            char_check=''
                                    sig_com=sig_com+'"'
                        
                        if string[0:5] == 'VAL_ ':
                            #ID сообщения
                            mes_id=''
                            i=5
                            char=string[i]
                            while char != ' ':
                                mes_id=mes_id+char
                                i=i+1
                                char=string[i]
                            mes_id=hex(int('0'+f"{int(mes_id):b}"[1:], 2))[2:].upper()
                            while len(mes_id) !=8:
                                mes_id='0'+mes_id
                            
                            if mes_id == message_id:
                                i=i+1
                                char=string[i]
                                sig_name=''
                                while char != ' ':
                                    sig_name=sig_name+char
                                    i=i+1
                                    char=string[i]
                                    
                                if sig_name == signal_name:
                                    result=re.findall(r'\d+ "[^"]*"', string)
                                    for i in result:
                                        i=re.sub(r' "', ': "', i)
                                        sig_value=sig_value+i+'\n'
                            
                        string=file.readline()
                    file.seek(ukz)             
                string=file.readline()
                #print(signal_startbit, signal_length)
                signals.append(DBC.signal(signal_name, int(signal_startbit), int(signal_length), float(signal_factor), float(signal_offset), signal_unit, sig_com, sig_value, signal_range))
            dbc.append(DBC.message(message_id, message_name, signals))              
        string=file.readline()
    return dbc

def new_excel(pathExcel, pathDBC, sheetname):
    dbc=DBC_open(pathDBC)
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title=sheetname
    
    font=Font('Calibri', 11, bold=True)
    align=Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
    border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cel=sheet.cell(1, 2, 'Параметр')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 3, 'Название CAN-сообщения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 4, 'ID')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 5, 'Сообщение')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 6, 'Start bit')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 7, 'Длинна, бит')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 8, 'Coeff.')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 9, 'Shift')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 10, 'Пределы измерения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 11, 'Единицы измерения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 12, 'Период, мс')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 13, 'Тип параметра')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 14, 'Расшифровка')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 15, 'Примечание')
    cel.font=font
    cel.border=border
    cel.alignment=align
    
    count=2
    
    flag='white'
    for i in dbc:
        for j in i.signals:
            if sheet.cell(count-1, 3).value != i.name:
                if flag=='white':
                    color=Color(rgb=None, indexed=None, auto=None, theme=4, tint=0.5999938962981048, type='theme')
                    fill=PatternFill(patternType='solid', fgColor=color)
                    flag='lightblue'
                elif flag=='lightblue':
                    color=Color(rgb='FFFFFFFF', indexed=None, auto=None, theme=None, tint=0.0, type='rgb')
                    fill=PatternFill(patternType='solid', fgColor=color)
                    flag='white'
                
            #sheet.write(count, 2, 'Параметр')
            sheet.cell(count, 1).fill=fill
            sheet.cell(count, 1).border=border
            sheet.cell(count, 2).fill=fill
            sheet.cell(count, 2).border=border
            sheet.cell(count, 3, i.name).alignment=align
            sheet.cell(count, 3).fill=fill
            sheet.cell(count, 3).border=border
            sheet.cell(count, 4, i.ID).alignment=align
            sheet.cell(count, 4).fill=fill
            sheet.cell(count, 4).border=border
            sheet.cell(count, 5, j.name).alignment=align
            sheet.cell(count, 5).fill=fill
            sheet.cell(count, 5).border=border
            sheet.cell(count, 6, j.startbit).alignment=align
            sheet.cell(count, 6).fill=fill
            sheet.cell(count, 6).border=border
            sheet.cell(count, 7, j.length).alignment=align
            sheet.cell(count, 7).fill=fill
            sheet.cell(count, 7).border=border
            sheet.cell(count, 8, j.factor).alignment=align
            sheet.cell(count, 8).fill=fill
            sheet.cell(count, 8).border=border
            sheet.cell(count, 9, j.offset).alignment=align
            sheet.cell(count, 9).fill=fill
            sheet.cell(count, 9).border=border
            sheet.cell(count, 10, j.sig_range).alignment=align
            sheet.cell(count, 10).fill=fill
            sheet.cell(count, 10).border=border
            sheet.cell(count, 11, j.unit).alignment=align
            sheet.cell(count, 11).fill=fill
            sheet.cell(count, 11).border=border
            sheet.cell(count, 12).fill=fill
            sheet.cell(count, 12).border=border
            sheet.cell(count, 13).fill=fill
            sheet.cell(count, 13).border=border
            #sheet.write(count, 12, 'Период, мс')
            if j.unit !='':
                sheet.cell(count, 13, 'Непрерывный')
            if j.sig_values !='':
                sheet.cell(count, 13, 'Дискретный')
            sheet.cell(count, 14, j.sig_values).alignment=align
            sheet.cell(count, 14).border=border
            sheet.cell(count, 14).fill=fill
            sheet.cell(count, 15, j.sig_com).alignment=align
            sheet.cell(count, 15).fill=fill
            sheet.cell(count, 15).border=border
            count +=1
    wb.save(pathExcel)
    
    wb=openpyxl.load_workbook(pathExcel)
    sheet=wb.active
    
    sheet.column_dimensions['A'].width=1.86+0.71
    sheet.column_dimensions['B'].width=30.71+0.71
    sheet.column_dimensions['C'].width=37.43+0.71
    sheet.column_dimensions['D'].width=9+0.71
    sheet.column_dimensions['E'].width=38.14+0.71
    sheet.column_dimensions['F'].width=7.71+0.71
    sheet.column_dimensions['G'].width=7.86+0.71
    sheet.column_dimensions['H'].width=10.29+0.71
    sheet.column_dimensions['I'].width=6+0.71
    sheet.column_dimensions['J'].width=19.57+0.71
    sheet.column_dimensions['K'].width=10.43+0.71
    sheet.column_dimensions['L'].width=10.57+0.71
    sheet.column_dimensions['M'].width=14+0.71
    sheet.column_dimensions['N'].width=34+0.71
    sheet.column_dimensions['O'].width=50+0.71
    wb.save(pathExcel)    

def add_sheet(pathExcel, pathDBC, sheetname):
    dbc=DBC_open(pathDBC)
    wb=openpyxl.load_workbook(pathExcel, data_only=True)
    path_old=pathExcel[:-5]+'_old'+pathExcel[-5:]
    wb.save(path_old)
    sheet=wb.create_sheet(sheetname)
    #sheet.title=sheetname
    font=Font('Calibri', 11, bold=True)
    align=Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
    border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cel=sheet.cell(1, 2, 'Параметр')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 3, 'Название CAN-сообщения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 4, 'ID')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 5, 'Сообщение')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 6, 'Start bit')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 7, 'Длинна, бит')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 8, 'Coeff.')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 9, 'Shift')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 10, 'Пределы измерения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 11, 'Единицы измерения')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 12, 'Период, мс')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 13, 'Тип параметра')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 14, 'Расшифровка')
    cel.font=font
    cel.border=border
    cel.alignment=align
    cel=sheet.cell(1, 15, 'Примечание')
    cel.font=font
    cel.border=border
    cel.alignment=align
    
    count=2
    
    flag='white'
    for i in dbc:
        for j in i.signals:
            if sheet.cell(count-1, 3).value != i.name:
                if flag=='white':
                    color=Color(rgb=None, indexed=None, auto=None, theme=4, tint=0.5999938962981048, type='theme')
                    fill=PatternFill(patternType='solid', fgColor=color)
                    flag='lightblue'
                elif flag=='lightblue':
                    color=Color(rgb='FFFFFFFF', indexed=None, auto=None, theme=None, tint=0.0, type='rgb')
                    fill=PatternFill(patternType='solid', fgColor=color)
                    flag='white'
                
            #sheet.write(count, 2, 'Параметр')
            sheet.cell(count, 1).fill=fill
            sheet.cell(count, 1).border=border
            sheet.cell(count, 2).fill=fill
            sheet.cell(count, 2).border=border
            sheet.cell(count, 3, i.name).alignment=align
            sheet.cell(count, 3).fill=fill
            sheet.cell(count, 3).border=border
            sheet.cell(count, 4, i.ID).alignment=align
            sheet.cell(count, 4).fill=fill
            sheet.cell(count, 4).border=border
            sheet.cell(count, 5, j.name).alignment=align
            sheet.cell(count, 5).fill=fill
            sheet.cell(count, 5).border=border
            sheet.cell(count, 6, j.startbit).alignment=align
            sheet.cell(count, 6).fill=fill
            sheet.cell(count, 6).border=border
            sheet.cell(count, 7, j.length).alignment=align
            sheet.cell(count, 7).fill=fill
            sheet.cell(count, 7).border=border
            sheet.cell(count, 8, j.factor).alignment=align
            sheet.cell(count, 8).fill=fill
            sheet.cell(count, 8).border=border
            sheet.cell(count, 9, j.offset).alignment=align
            sheet.cell(count, 9).fill=fill
            sheet.cell(count, 9).border=border
            sheet.cell(count, 10, j.sig_range).alignment=align
            sheet.cell(count, 10).fill=fill
            sheet.cell(count, 10).border=border
            sheet.cell(count, 11, j.unit).alignment=align
            sheet.cell(count, 11).fill=fill
            sheet.cell(count, 11).border=border
            sheet.cell(count, 12).fill=fill
            sheet.cell(count, 12).border=border
            sheet.cell(count, 13).fill=fill
            sheet.cell(count, 13).border=border
            #sheet.write(count, 12, 'Период, мс')
            if j.unit !='':
                sheet.cell(count, 13, 'Непрерывный')
            if j.sig_values !='':
                sheet.cell(count, 13, 'Дискретный')
            sheet.cell(count, 14, j.sig_values).alignment=align
            sheet.cell(count, 14).border=border
            sheet.cell(count, 14).fill=fill
            sheet.cell(count, 15, j.sig_com).alignment=align
            sheet.cell(count, 15).fill=fill
            sheet.cell(count, 15).border=border
            count +=1
    wb.save(pathExcel)
    
    wb=openpyxl.load_workbook(pathExcel)
    sheet=wb[sheetname]
    
    sheet.column_dimensions['A'].width=1.86+0.71
    sheet.column_dimensions['B'].width=30.71+0.71
    sheet.column_dimensions['C'].width=37.43+0.71
    sheet.column_dimensions['D'].width=9+0.71
    sheet.column_dimensions['E'].width=38.14+0.71
    sheet.column_dimensions['F'].width=7.71+0.71
    sheet.column_dimensions['G'].width=7.86+0.71
    sheet.column_dimensions['H'].width=10.29+0.71
    sheet.column_dimensions['I'].width=6+0.71
    sheet.column_dimensions['J'].width=19.57+0.71
    sheet.column_dimensions['K'].width=10.43+0.71
    sheet.column_dimensions['L'].width=10.57+0.71
    sheet.column_dimensions['M'].width=14+0.71
    sheet.column_dimensions['N'].width=34+0.71
    sheet.column_dimensions['O'].width=50+0.71
    
    wb.save(pathExcel)   
    
def add_sig(pathExcel, pathDBC, sheetname):
    dbc=DBC_open(pathDBC)
    wb=openpyxl.load_workbook(pathExcel, data_only=True)
    path_old=pathExcel[:-5]+'_old'+pathExcel[-5:]
    wb.save(path_old)
    sheet=wb[sheetname]
    
    
    for i in dbc:
        count=2
        flag=0
        while sheet.cell(count, 3).value != None:
            if i.name == sheet.cell(count, 3).value:
                flag=1
                break
            count +=1
        if flag==0:
            c=sheet.cell(count, 1).fill.fgColor.tint
            if c == 0.0:
                flag_c='white'
            elif c==0.5999938962981048:
                flag_c='lightblue'
            for j in i.signals:
                if sheet.cell(count-1, 3).value != i.name:
                    if flag_c=='white':
                        color=Color(rgb=None, indexed=None, auto=None, theme=4, tint=0.5999938962981048, type='theme')
                        fill=PatternFill(patternType='solid', fgColor=color)
                        flag_c='lightblue'
                    elif flag_c=='lightblue':
                        color=Color(rgb='FFFFFFFF', indexed=None, auto=None, theme=None, tint=0.0, type='rgb')
                        fill=PatternFill(patternType='solid', fgColor=color)
                        flag_c='white'
                align=Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
                border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))    
                #sheet.write(count, 2, 'Параметр')
                sheet.cell(count, 1).fill=fill
                sheet.cell(count, 1).border=border
                sheet.cell(count, 2).fill=fill
                sheet.cell(count, 2).border=border
                sheet.cell(count, 3, i.name).alignment=align
                sheet.cell(count, 3).fill=fill
                sheet.cell(count, 3).border=border
                sheet.cell(count, 4, i.ID).alignment=align
                sheet.cell(count, 4).fill=fill
                sheet.cell(count, 4).border=border
                sheet.cell(count, 5, j.name).alignment=align
                sheet.cell(count, 5).fill=fill
                sheet.cell(count, 5).border=border
                sheet.cell(count, 6, j.startbit).alignment=align
                sheet.cell(count, 6).fill=fill
                sheet.cell(count, 6).border=border
                sheet.cell(count, 7, j.length).alignment=align
                sheet.cell(count, 7).fill=fill
                sheet.cell(count, 7).border=border
                sheet.cell(count, 8, j.factor).alignment=align
                sheet.cell(count, 8).fill=fill
                sheet.cell(count, 8).border=border
                sheet.cell(count, 9, j.offset).alignment=align
                sheet.cell(count, 9).fill=fill
                sheet.cell(count, 9).border=border
                sheet.cell(count, 10, j.sig_range).alignment=align
                sheet.cell(count, 10).fill=fill
                sheet.cell(count, 10).border=border
                sheet.cell(count, 11, j.unit).alignment=align
                sheet.cell(count, 11).fill=fill
                sheet.cell(count, 11).border=border
                sheet.cell(count, 12).fill=fill
                sheet.cell(count, 12).border=border
                sheet.cell(count, 13).fill=fill
                sheet.cell(count, 13).border=border
                #sheet.write(count, 12, 'Период, мс')
                if j.unit !='':
                    sheet.cell(count, 13, 'Непрерывный')
                if j.sig_values !='':
                    sheet.cell(count, 13, 'Дискретный')
                sheet.cell(count, 14, j.sig_values).alignment=align
                sheet.cell(count, 14).border=border
                sheet.cell(count, 14).fill=fill
                sheet.cell(count, 15, j.sig_com).alignment=align
                sheet.cell(count, 15).fill=fill
                sheet.cell(count, 15).border=border
                count +=1
        if flag==1:
            #fill=sheet.cell(count, 1).fill
            align=Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
            border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 
            for j in i.signals:
                flag_s=0
                while i.name == sheet.cell(count, 3).value:
                    if j.name == sheet.cell(count, 5).value:
                        flag_s=1
                        break
                    count += 1
                if flag_s == 0:
                    sheet.insert_rows(count)
                    #sheet.cell(count, 1).fill=fill
                    sheet.cell(count, 1).border=border
                    #sheet.cell(count, 2).fill=fill
                    sheet.cell(count, 2).border=border
                    sheet.cell(count, 3, i.name).alignment=align
                    #sheet.cell(count, 3).fill=fill
                    sheet.cell(count, 3).border=border
                    sheet.cell(count, 4, i.ID).alignment=align
                    #sheet.cell(count, 4).fill=fill
                    sheet.cell(count, 4).border=border
                    sheet.cell(count, 5, j.name).alignment=align
                    #sheet.cell(count, 5).fill=fill
                    sheet.cell(count, 5).border=border
                    sheet.cell(count, 6, j.startbit).alignment=align
                    #sheet.cell(count, 6).fill=fill
                    sheet.cell(count, 6).border=border
                    sheet.cell(count, 7, j.length).alignment=align
                    #sheet.cell(count, 7).fill=fill
                    sheet.cell(count, 7).border=border
                    sheet.cell(count, 8, j.factor).alignment=align
                    #sheet.cell(count, 8).fill=fill
                    sheet.cell(count, 8).border=border
                    sheet.cell(count, 9, j.offset).alignment=align
                    #sheet.cell(count, 9).fill=fill
                    sheet.cell(count, 9).border=border
                    sheet.cell(count, 10, j.sig_range).alignment=align
                    #sheet.cell(count, 10).fill=fill
                    sheet.cell(count, 10).border=border
                    sheet.cell(count, 11, j.unit).alignment=align
                    #sheet.cell(count, 11).fill=fill
                    sheet.cell(count, 11).border=border
                    #sheet.cell(count, 12).fill=fill
                    sheet.cell(count, 12).border=border
                    #sheet.cell(count, 13).fill=fill
                    sheet.cell(count, 13).border=border
                    #sheet.write(count, 12, 'Период, мс')
                    if j.unit !='':
                        sheet.cell(count, 13, 'Непрерывный')
                    if j.sig_values !='':
                        sheet.cell(count, 13, 'Дискретный')
                    sheet.cell(count, 14, j.sig_values).alignment=align
                    sheet.cell(count, 14).border=border
                    #sheet.cell(count, 14).fill=fill
                    sheet.cell(count, 15, j.sig_com).alignment=align
                    #sheet.cell(count, 15).fill=fill
                    sheet.cell(count, 15).border=border                  
                if flag_s == 1:
                    sheet.cell(count, 6, j.startbit).alignment=align
                    #sheet.cell(count, 6).fill=fill
                    sheet.cell(count, 6).border=border
                    sheet.cell(count, 7, j.length).alignment=align
                    #sheet.cell(count, 7).fill=fill
                    sheet.cell(count, 7).border=border
                    sheet.cell(count, 8, j.factor).alignment=align
                    #sheet.cell(count, 8).fill=fill
                    sheet.cell(count, 8).border=border
                    sheet.cell(count, 9, j.offset).alignment=align
                    #sheet.cell(count, 9).fill=fill
                    sheet.cell(count, 9).border=border
                    sheet.cell(count, 10, j.sig_range).alignment=align
                    #sheet.cell(count, 10).fill=fill
                    sheet.cell(count, 10).border=border
                    sheet.cell(count, 11, j.unit).alignment=align
                    #sheet.cell(count, 11).fill=fill
                    sheet.cell(count, 11).border=border
                    #sheet.cell(count, 12).fill=fill
                    sheet.cell(count, 12).border=border
                    #sheet.cell(count, 13).fill=fill
                    sheet.cell(count, 13).border=border
                    #sheet.write(count, 12, 'Период, мс')
                    if j.unit !='':
                        sheet.cell(count, 13, 'Непрерывный')
                    if j.sig_values !='':
                        sheet.cell(count, 13, 'Дискретный')
                    sheet.cell(count, 14, j.sig_values).alignment=align
                    sheet.cell(count, 14).border=border
                    #sheet.cell(count, 14).fill=fill
                    sheet.cell(count, 15, j.sig_com).alignment=align
                    #sheet.cell(count, 15).fill=fill
                    sheet.cell(count, 15).border=border
                    
    wb.save(pathExcel)  
if __name__ == '__main__':
    app=QtWidgets.QApplication(sys.argv)
    application = Main()
    application.show()
    sys.exit(app.exec_())
